# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import joblib
import json
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
import traceback # 用于打印详细错误信息
import collections # 用于检查重复列名

# --- 配置参数 ---
MODEL_FILE = "optimized_model_v2_ensemble.pkl"  # 保存的模型文件
META_FILE = "optimized_meta_data_v2.json"  # 保存的元数据文件
TEST_DATA = "测试用的一条数据.xlsx"  # 测试数据文件
OUTPUT_FILE = "单条数据优化结果.xlsx"  # 优化结果输出文件

# --- 列名常量 (请根据你的Excel文件调整) ---
ORIGINAL_DATE_COLUMN = "测量日期" # 指定原始的日期列名
ORIGINAL_PUE_COLUMN = "PUE值"   # 指定原始的PUE列名 (将被从输出中移除)

# --- 加载模型和元数据 ---
def load_saved_models():
    try:
        models, scalers, feature_lists = joblib.load(MODEL_FILE)
        with open(META_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)
        # 确保config中的列表是list类型
        config['variable_cols'] = list(config.get('variable_cols', []))
        config['integer_cols'] = list(config.get('integer_cols', []))
        # 检查必要的键是否存在
        required_keys = ['variable_cols', 'target_energy_col', 'variable_bounds', 'integer_cols']
        for key in required_keys:
            if key not in config:
                raise KeyError(f"元数据文件 '{META_FILE}' 缺少必要的键: '{key}'")
        if 'temp1' not in models or 'temp2' not in models or 'energy' not in models:
             raise KeyError(f"模型文件 '{MODEL_FILE}' 缺少必要的模型: 'temp1', 'temp2', 或 'energy'")
        if 'energy' not in scalers:
             raise KeyError(f"模型文件 '{MODEL_FILE}' 缺少必要的Scaler: 'energy'")
        if 'temp1' not in feature_lists or 'temp2' not in feature_lists or 'energy' not in feature_lists:
             raise KeyError(f"模型文件 '{MODEL_FILE}' 缺少必要的特征列表: 'temp1', 'temp2', 或 'energy'")

        return models, scalers, feature_lists, config
    except FileNotFoundError as e:
        print(f"错误: 文件未找到 - {e}. 请确保模型和元数据文件存在于脚本所在目录或指定路径。")
        raise
    except Exception as e:
        print(f"加载模型或元数据失败: {e}")
        raise

# --- PSO 优化器类 ---
class ALF_PSO:
    """自适应学习因子PSO优化器 (处理混合整数变量)"""

    def __init__(self, models, scalers, feature_lists, data_row_series, config):
        # --- 模型和Scaler ---
        self.model_temp1 = models['temp1']
        self.model_temp2 = models['temp2']
        self.model_energy = models['energy']
        self.energy_scaler = scalers['energy']

        # --- 特征列表 ---
        self.temp1_features = list(feature_lists['temp1'])
        self.temp2_features = list(feature_lists['temp2'])
        self.energy_features = list(feature_lists['energy'])

        # --- 配置和数据 ---
        self.original_row = data_row_series.copy() # 使用包含原始列名的Series
        self.variable_cols = list(config['variable_cols'])
        self.target_energy_col = config['target_energy_col']
        self.variable_bounds_dict = config['variable_bounds']
        self.integer_cols = list(config.get('integer_cols', []))

        # --- 检查数据完整性 (使用原始列名检查传入的data_row_series) ---
        missing_in_row = [col for col in self.variable_cols if col not in self.original_row.index]
        if missing_in_row:
            raise ValueError(f"初始化ALF_PSO时, data_row_series (应含原始列名) 缺少变量列: {missing_in_row}")
        required_features = set(self.temp1_features) | set(self.temp2_features) | set(self.energy_features)
        missing_features_in_row = [f for f in required_features if f not in self.original_row.index]
        if missing_features_in_row:
             raise ValueError(f"初始化ALF_PSO时, data_row_series (应含原始列名) 缺少模型所需特征: {missing_features_in_row}")
        if self.target_energy_col not in self.original_row.index:
             raise ValueError(f"初始化ALF_PSO时, data_row_series (应含原始列名) 缺少目标能耗列: {self.target_energy_col}")

        # --- 提取常量值 (使用原始列名) ---
        self.const_temp_values = {
            col: self.original_row[col] for col in self.temp1_features
            if col not in self.variable_cols and col in self.original_row.index
        }
        self.const_energy_values = {
             col: self.original_row[col] for col in self.energy_features
             if col not in self.variable_cols and col in self.original_row.index
        }

        # --- 优化状态 ---
        self.original_energy = self.original_row[self.target_energy_col]
        self.best_energy = np.inf
        self.best_params_vec = None
        self.dim = len(self.variable_cols)

        # --- 边界和整数索引 ---
        try:
            self.lb = np.array([self.variable_bounds_dict[col][0] for col in self.variable_cols])
            self.ub = np.array([self.variable_bounds_dict[col][1] for col in self.variable_cols])
        except KeyError as e:
             raise KeyError(f"变量 '{e}' 在 variable_bounds_dict 中没有定义边界。请检查元数据文件。")
        except TypeError as e:
             raise TypeError(f"variable_bounds_dict 结构错误，无法获取边界: {e}。请检查元数据文件。")


        self.integer_indices = [i for i, col in enumerate(self.variable_cols)
                                if col in self.integer_cols]

    def _round_integer_vars(self, x_vec):
        """将向量中指定索引的值四舍五入到0或1"""
        if not self.integer_indices:
            return x_vec
        x_rounded = x_vec.copy()
        for idx in self.integer_indices:
            x_rounded[idx] = 1.0 if x_rounded[idx] >= 0.5 else 0.0
        return x_rounded

    def _create_input_df(self, x_vec, feature_list, const_values):
        """根据当前参数向量创建模型输入的DataFrame"""
        x_processed = self._round_integer_vars(x_vec)
        input_data = {col: x_processed[i] for i, col in enumerate(self.variable_cols)}
        input_data.update(const_values)

        missing_features = [f for f in feature_list if f not in input_data]
        for mf in missing_features:
            if mf in self.original_row.index:
                input_data[mf] = self.original_row[mf]
            else:
                print(f"错误: 模型特征 '{mf}' 在原始数据行中也找不到。")
                return None

        try:
            all_features_dict = {f: np.nan for f in feature_list}
            all_features_dict.update(input_data)
            input_df = pd.DataFrame([all_features_dict])[feature_list]
            if input_df.isnull().values.any():
                 print(f"警告: 为模型创建的输入DataFrame包含NaN值。 Features: {feature_list}")
                 # print(input_df[input_df.isnull().any(axis=1)])
                 return None
            return input_df
        except Exception as e:
            print(f"创建模型输入DataFrame时出错: {e}")
            # print(f"Feature List: {feature_list}")
            # print(f"Input Data Keys: {list(input_data.keys())}")
            return None

    def objective_function(self, x_vec):
        """PSO的目标函数：预测能耗并施加约束惩罚"""
        try:
            # 0. 边界检查
            if np.any(x_vec < self.lb) or np.any(x_vec > self.ub):
                return np.inf

            # 1. 预测温度
            temp1_input_df = self._create_input_df(x_vec, self.temp1_features, self.const_temp_values)
            temp2_input_df = self._create_input_df(x_vec, self.temp2_features, self.const_temp_values)

            if temp1_input_df is None or temp2_input_df is None:
                return np.inf

            pred_temp1 = self.model_temp1.predict(temp1_input_df)[0]
            pred_temp2 = self.model_temp2.predict(temp2_input_df)[0]

            # 2. 温度约束检查和惩罚
            temp_constraint_penalty = 0
            max_temp1_limit = 21.55
            max_temp2_limit = 20.91
            if pred_temp1 > max_temp1_limit:
                temp_constraint_penalty += (pred_temp1 - max_temp1_limit + 1)**2 * 1e6
            if pred_temp2 > max_temp2_limit:
                temp_constraint_penalty += (pred_temp2 - max_temp2_limit + 1)**2 * 1e6

            if temp_constraint_penalty > 0:
                return temp_constraint_penalty + 1e9 # 比正常能耗大得多

            # 3. 预测能耗 (只有在温度约束满足时才进行)
            energy_input_df = self._create_input_df(x_vec, self.energy_features, self.const_energy_values)
            if energy_input_df is None:
                return np.inf

            # 缩放能耗输入
            try:
                energy_input_scaled = self.energy_scaler.transform(energy_input_df)
            except ValueError as ve: # 捕获维度或类型不匹配等错误
                 print(f"错误: 缩放能耗输入时发生值错误: {ve}")
                 print(f"Energy Input DF shape: {energy_input_df.shape}, columns: {energy_input_df.columns}")
                 # 尝试打印 Scaler 期望的特征数量或名称（如果可用）
                 if hasattr(self.energy_scaler, 'n_features_in_'):
                      print(f"Scaler expected features: {self.energy_scaler.n_features_in_}")
                 if hasattr(self.energy_scaler, 'feature_names_in_'):
                      print(f"Scaler expected feature names: {self.energy_scaler.feature_names_in_}")
                 return np.inf # 返回无效
            except Exception as e:
                print(f"缩放能耗输入时出错: {e}")
                return np.inf

            pred_energy = self.model_energy.predict(energy_input_scaled)[0]

            # 4. 计算最终目标值
            if pred_energy >= self.original_energy:
                 return self.original_energy + abs(pred_energy - self.original_energy) + 1e-6
            else:
                 return pred_energy

        except Exception as e:
            # print(f"目标函数计算中发生异常: {e}")
            # traceback.print_exc() # 调试时可以取消注释
            return np.inf

    def optimize(self, swarmsize=27, maxiter=10, w=0.729, c1_range=(0.5, 2.5), c2_range=(0.5, 2.5)):
        """执行ALF-PSO优化过程"""
        c1_min, c1_max = c1_range
        c2_min, c2_max = c2_range

        swarm_pos = np.random.uniform(self.lb, self.ub, (swarmsize, self.dim))
        swarm_pos = np.array([self._round_integer_vars(p) for p in swarm_pos])
        vel_range = (self.ub - self.lb) * 0.1
        swarm_vel = np.random.uniform(-vel_range, vel_range, (swarmsize, self.dim))
        swarm_pbest_pos = swarm_pos.copy()
        swarm_pbest_val = np.array([self.objective_function(p) for p in swarm_pos])

        if np.all(np.isinf(swarm_pbest_val)):
             print("警告: 所有初始粒子的目标函数值均为无穷大。检查参数/约束。")
             gbest_idx = 0; gbest_val = np.inf
        else:
             valid_indices = np.where(np.isfinite(swarm_pbest_val))[0]
             if len(valid_indices) == 0: gbest_idx = 0; gbest_val = np.inf
             else:
                  gbest_idx = valid_indices[np.nanargmin(swarm_pbest_val[valid_indices])]
                  gbest_val = swarm_pbest_val[gbest_idx]
        gbest_pos = swarm_pbest_pos[gbest_idx].copy()

        for iter_num in range(maxiter):
            c1 = c1_max - (c1_max - c1_min) * iter_num / maxiter
            c2 = c2_min + (c2_max - c2_min) * iter_num / maxiter
            r1 = np.random.random((swarmsize, self.dim))
            r2 = np.random.random((swarmsize, self.dim))

            swarm_vel = w * swarm_vel + \
                        c1 * r1 * (swarm_pbest_pos - swarm_pos) + \
                        c2 * r2 * (gbest_pos - swarm_pos)
            swarm_pos = swarm_pos + swarm_vel
            swarm_pos = np.clip(swarm_pos, self.lb, self.ub)
            swarm_pos = np.array([self._round_integer_vars(p) for p in swarm_pos])
            current_val = np.array([self.objective_function(p) for p in swarm_pos])

            update_mask = current_val < swarm_pbest_val
            swarm_pbest_pos[update_mask] = swarm_pos[update_mask]
            swarm_pbest_val[update_mask] = current_val[update_mask]

            if np.any(update_mask):
                if np.all(np.isinf(swarm_pbest_val)): iter_best_val = np.inf
                else:
                    valid_indices = np.where(np.isfinite(swarm_pbest_val))[0]
                    if len(valid_indices) > 0:
                         iter_best_idx = valid_indices[np.nanargmin(swarm_pbest_val[valid_indices])]
                         iter_best_val = swarm_pbest_val[iter_best_idx]
                         if iter_best_val < gbest_val:
                              gbest_pos = swarm_pbest_pos[iter_best_idx].copy()
                              gbest_val = iter_best_val
                    else: iter_best_val = np.inf

            # print(f"Iter {iter_num+1}/{maxiter}, Best Val: {gbest_val:.4f}")

        if np.isfinite(gbest_val) and gbest_val < self.original_energy - 1e-9 :
            self.best_energy = gbest_val
            self.best_params_vec = gbest_pos
            print(f"优化找到更优解: {self.best_energy:.4f} (原始能耗: {self.original_energy:.4f})")
            return self.best_params_vec
        else:
            print(f"未找到优于原始能耗 ({self.original_energy:.4f}) 的解。最优尝试值: {gbest_val:.4f}")
            return None

# --- 保存结果函数 ---
def save_optimized_result(original_data_row,    # 包含清洗后列名的原始数据行(Series)
                          optimized_params_vec, # 优化后的参数值(numpy array)，如果无优化则为None
                          optimized_energy,     # 优化后的能耗值(float)
                          variable_cols,        # 包含清洗后列名的变量列表(list)
                          target_energy_col,    # 清洗后的目标能耗列名(str)
                          date_col_name,        # 清洗后的日期列名(str)，可能为None
                          col_to_drop_cleaned,  # 清洗后的要删除的列名(str)，可能为None
                          output_file):
    """
    保存优化结果到Excel。
    - 如果优化成功，用优化值更新变量列和目标能耗列。
    - 从结果中删除指定的列 (col_to_drop_cleaned)。
    - 将发生变化的变量列（非日期列）和变化了的目标能耗列的字体标红。
    """
    result_df = pd.DataFrame([original_data_row]) # 创建单行DataFrame用于操作
    original_energy = original_data_row[target_energy_col]

    optimization_improved = False

    # 检查优化是否成功且找到了更优解
    if optimized_params_vec is not None and np.isfinite(optimized_energy) and optimized_energy < original_energy - 1e-9:
        optimization_improved = True
        print(f"准备更新结果: 优化能耗 {optimized_energy:.4f} < 原始能耗 {original_energy:.4f}")
        current_index = result_df.index[0]
        # 更新变量列值
        for i, col_name in enumerate(variable_cols):
             if col_name in result_df.columns:
                 result_df.at[current_index, col_name] = optimized_params_vec[i]
             else:
                 print(f"警告: 变量列 '{col_name}' 在结果DataFrame中不存在，无法更新。")
        # 更新目标能耗列值
        if target_energy_col in result_df.columns:
            result_df.at[current_index, target_energy_col] = optimized_energy
        else:
             print(f"警告: 目标能耗列 '{target_energy_col}' 在结果DataFrame中不存在，无法更新。")
    else:
        print(f"不更新结果: 优化未成功或未找到更优解 (Optimized: {optimized_energy}, Original: {original_energy})")

    # --- 在保存前删除指定列 ---
    if col_to_drop_cleaned and col_to_drop_cleaned in result_df.columns:
        try:
            result_df.drop(columns=[col_to_drop_cleaned], inplace=True)
            print(f"已从输出结果DataFrame中移除列: '{col_to_drop_cleaned}'")
        except Exception as drop_e:
             print(f"警告: 尝试移除列 '{col_to_drop_cleaned}' 时出错: {drop_e}")
    elif col_to_drop_cleaned:
         print(f"警告: 想要移除的列 '{col_to_drop_cleaned}' 不在结果DataFrame的列中，无法移除。")

    # --- 获取实际要保存的列标题 (删除列之后) ---
    headers = list(result_df.columns)
    if not headers:
         print("错误: 删除列后DataFrame为空，无法保存。")
         return

    # 3. 保存处理后的DataFrame到Excel文件
    try:
        result_df.to_excel(output_file, index=False, engine='openpyxl')
    except Exception as e:
        print(f"保存Excel文件 '{output_file}' 时出错: {e}")
        return

    # 4. 应用红色字体格式化 (基于已保存的文件结构)
    try:
        wb = load_workbook(output_file)
        ws = wb.active
        red_font = Font(color="FF0000")
        changed_cells_count = 0

        # 检查并标记变化的变量列 (排除日期列，且列必须存在于headers中)
        if optimized_params_vec is not None:
            for i, col_name in enumerate(variable_cols):
                if col_name == date_col_name: continue # 跳过日期列
                if col_name == col_to_drop_cleaned: continue # 跳过已删除的列
                if col_name not in headers: continue # 跳过不存在于最终结果的列

                # 检查原始数据行中是否有此列，以进行比较
                if col_name not in original_data_row.index:
                     print(f"信息: 变量列 '{col_name}' 不在原始数据中，无法比较是否变化。")
                     continue

                original_val = original_data_row[col_name]
                optimized_val = optimized_params_vec[i]

                is_different = False
                try:
                    # 比较，注意类型转换和浮点数精度
                    if isinstance(original_val, (int, float, np.number)) and isinstance(optimized_val, (int, float, np.number)):
                        if not np.isclose(float(original_val), float(optimized_val), rtol=1e-5, atol=1e-8):
                            is_different = True
                    # bool/int 0/1 的情况，也可能需要特殊处理，这里用简单比较
                    elif type(original_val) is type(optimized_val) and original_val != optimized_val:
                         is_different = True
                    # 处理类型不同的情况，可能也算变化 (例如 None 变为 0)
                    elif type(original_val) is not type(optimized_val):
                         # 可能需要更复杂的逻辑判断这是否算“有效”变化
                         # 暂时认为类型不同就是变化，除非两者都是None或NaN等效
                         if not (pd.isna(original_val) and pd.isna(optimized_val)):
                              is_different = True
                except TypeError as te:
                    print(f"警告: 比较列 '{col_name}' 时类型错误: {te}。原始: {original_val}, 优化: {optimized_val}。跳过。")
                    continue

                if is_different:
                    try:
                        col_idx_excel = headers.index(col_name) + 1
                        cell = ws.cell(row=2, column=col_idx_excel)
                        cell.font = red_font
                        changed_cells_count += 1
                    except ValueError: pass # 列不在headers中 (理论上前面已检查)
                    except Exception as cell_e: print(f"警告: 标记变量单元格(列'{col_name}')时出错: {cell_e}")

        # 检查并标记变化了的目标能耗列 (如果它没被删除)
        if target_energy_col in headers:
            current_energy_in_df = result_df.at[result_df.index[0], target_energy_col] # 获取最终值
            if not np.isclose(original_energy, current_energy_in_df, rtol=1e-5, atol=1e-8):
                 try:
                     col_idx_excel = headers.index(target_energy_col) + 1
                     cell = ws.cell(row=2, column=col_idx_excel)
                     cell.font = red_font
                     changed_cells_count += 1
                 except ValueError: pass
                 except Exception as cell_e: print(f"警告: 标记能耗单元格时出错: {cell_e}")

        # 保存格式化更改
        if changed_cells_count > 0:
            print(f"共标记了 {changed_cells_count} 个变化的单元格为红色。")
            wb.save(output_file)
        else:
            print("没有检测到需要标红的变化。")

    except Exception as e:
        print(f"加载或保存Excel文件 '{output_file}' 进行格式化时出错: {e}")
        traceback.print_exc()


# --- 列名清洗函数 ---
def clean_col_name(col_name):
    """清洗列名，移除非法字符"""
    cleaned = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9_]+', '_', str(col_name))
    cleaned = cleaned.strip('_')
    cleaned = re.sub(r'_+', '_', cleaned)
    return cleaned

# --- 主函数 ---
def main():
    try:
        # 1. 加载模型、Scaler和元数据
        print("加载模型和元数据...")
        models, scalers, feature_lists, config = load_saved_models()
        original_variable_cols = config['variable_cols']
        original_target_energy_col = config['target_energy_col']

        # 2. 加载测试数据
        print(f"加载测试数据 '{TEST_DATA}'...")
        try:
            test_data = pd.read_excel(TEST_DATA)
        except FileNotFoundError:
             print(f"错误: 测试数据文件 '{TEST_DATA}' 未找到。")
             return
        except Exception as read_e:
             print(f"错误: 读取测试数据文件 '{TEST_DATA}' 时出错: {read_e}")
             return


        if test_data.empty: print(f"错误: 测试数据文件 '{TEST_DATA}' 为空。"); return
        if len(test_data) > 1: print(f"警告: 测试数据文件 '{TEST_DATA}' 含多行，仅用第一行。")

        # --- 列名处理 ---
        original_cols_list = list(test_data.columns)
        cleaned_cols_map = {orig: clean_col_name(orig) for orig in original_cols_list}
        cleaned_cols_list = [cleaned_cols_map[orig] for orig in original_cols_list]

        # 检查清洗后是否有重复列名
        duplicates = [item for item, count in collections.Counter(cleaned_cols_list).items() if count > 1]
        if duplicates:
            print(f"警告: 清洗后的列名存在重复: {duplicates}。可能导致问题。")

        # --- 获取用于PSO的原始列名数据行 ---
        data_row_original_names = test_data.iloc[0].copy()
        required_original_cols_for_pso = set(original_variable_cols) | set(feature_lists['temp1']) | set(feature_lists['temp2']) | set(feature_lists['energy'])
        missing_original_cols = [col for col in required_original_cols_for_pso if col not in data_row_original_names.index]
        if missing_original_cols: raise ValueError(f"测试数据 '{TEST_DATA}' 缺少PSO所需的原始列: {missing_original_cols}。")
        if original_target_energy_col not in data_row_original_names.index: raise ValueError(f"测试数据 '{TEST_DATA}' 缺少原始目标能耗列: '{original_target_energy_col}'。")

        # --- 获取用于保存结果的清洗后列名数据行 ---
        test_data_cleaned = test_data.copy()
        test_data_cleaned.columns = cleaned_cols_list
        data_row_cleaned_names = test_data_cleaned.iloc[0].copy()

        # --- 确定清洗后的关键列名 ---
        try:
            cleaned_variable_cols = [cleaned_cols_map[orig] for orig in original_variable_cols]
            cleaned_target_energy_col = cleaned_cols_map[original_target_energy_col]
            cleaned_date_col_name = cleaned_cols_map.get(ORIGINAL_DATE_COLUMN)
            cleaned_pue_col_name = cleaned_cols_map.get(ORIGINAL_PUE_COLUMN) # 获取PUE清洗后的名字

            if not cleaned_date_col_name: print(f"警告: 原始日期列 '{ORIGINAL_DATE_COLUMN}' 未在测试数据中找到。")
            if not cleaned_pue_col_name: print(f"警告: 原始PUE列 '{ORIGINAL_PUE_COLUMN}' 未在测试数据中找到，无法从输出移除。")

        except KeyError as e:
             raise KeyError(f"配置中的列名 '{e}' 在测试数据 '{TEST_DATA}' 的原始列中找不到。")


        # 3. 创建优化器实例
        print("创建优化器 (使用原始列名)...")
        optimizer = ALF_PSO(
            models=models, scalers=scalers, feature_lists=feature_lists,
            data_row_series=data_row_original_names, config=config
        )

        # 4. 执行优化
        print("执行ALF-PSO优化...")
        optimized_params_vec = optimizer.optimize()
        optimized_energy_value = optimizer.best_energy

        # 5. 保存结果 (传入清洗后列名的数据和要删除的PUE列名)
        print(f"保存优化结果到 '{OUTPUT_FILE}' (使用清洗后列名)...")
        save_optimized_result(
            original_data_row=data_row_cleaned_names,
            optimized_params_vec=optimized_params_vec,
            optimized_energy=optimized_energy_value,
            variable_cols=cleaned_variable_cols,
            target_energy_col=cleaned_target_energy_col,
            date_col_name=cleaned_date_col_name,
            col_to_drop_cleaned=cleaned_pue_col_name, # 传递清洗后的PUE列名
            output_file=OUTPUT_FILE
        )

        # 6. 打印总结信息
        original_energy_val = data_row_cleaned_names[cleaned_target_energy_col]
        if optimized_params_vec is not None and np.isfinite(optimized_energy_value) and optimized_energy_value < original_energy_val - 1e-9:
            print("-" * 30)
            print(f"优化完成! 结果已更新并保存到 {OUTPUT_FILE}。")
            if cleaned_pue_col_name: print(f"列 '{cleaned_pue_col_name}' 已从输出中移除。")
            print(f"变化的参数和能耗已标记为红色 (日期列除外)。")
            print(f"原始能耗 ({cleaned_target_energy_col}): {original_energy_val:.4f} kW")
            print(f"优化后能耗 ({cleaned_target_energy_col}): {optimized_energy_value:.4f} kW")
            print(f"能耗降低: {original_energy_val - optimized_energy_value:.4f} kW")
            print("-" * 30)
        else:
            print("-" * 30)
            print(f"未找到更优的能耗解决方案。")
            print(f"结果文件 '{OUTPUT_FILE}' 已保存。")
            if cleaned_pue_col_name: print(f"列 '{cleaned_pue_col_name}' 已从输出中移除。")
            print(f"原始能耗 ({cleaned_target_energy_col}): {original_energy_val:.4f} kW")
            print(f"优化器找到的最佳值为: {optimized_energy_value:.4f} kW")
            print("无单元格被标红（或仅标记了未改善能耗的变量变化）。")
            print("-" * 30)

    except FileNotFoundError as e: print(f"错误: 文件未找到 - {e}")
    except KeyError as e: print(f"错误: 缺少键或列名 - {e}。检查元数据、配置或Excel列名。")
    except ValueError as e: print(f"错误: 数据或配置值无效 - {e}")
    except Exception as e:
        print(f"程序运行过程中发生未预料的错误: {e}")
        print("-" * 60); print("详细错误追踪信息:"); traceback.print_exc(); print("-" * 60)

if __name__ == "__main__":
    main()