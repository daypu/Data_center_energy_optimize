import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# 加载原始数据和更新后的数据
original_data_path = 'PUE数据汇总.xlsx'
updated_data_path = 'PSO_调优后的参数_v2.xlsx'

original_data = pd.read_excel(original_data_path)
updated_data = pd.read_excel(updated_data_path)

# 需要对比的列名
columns_to_compare = [
    '冷机冷却侧出水压力PIT_1_CWS_1_压力表_实际值',
    '冷机冷却侧出水压力PIT_2_CWS_1_压力表_实际值',
    '冷机冷冻侧出水压力PIT_1_CWHS_1_压力表_实际值',
    '冷机冷冻侧出水压力PIT_2_CWHS_1_压力表_实际值',
    '冷机冷冻侧出水温度TIT_1_CWHS_1_温度计_实际值',
    '冷机冷冻侧出水温度TIT_2_CWHS_1_温度计_实际值',
    '冷机冷却侧出水温度TIT_1_CWS_1_温度计_实际值',
    '冷机冷却侧出水温度TIT_2_CWS_1_温度计_实际值',
    'CV1_1_比例阀门_开度反馈显示',
    'CV1_3_比例阀门_开度反馈显示',
    'CV10_1_比例阀门_开度反馈显示',
    'CV10_2_比例阀门_开度反馈显示',
    'CV11_1_比例阀门_开度反馈显示',
    'CV11_2_比例阀门_开度反馈显示',
    'CV14_1_比例阀门_开度反馈显示',
    'CV14_2_比例阀门_开度反馈显示',
    'CV15_1_比例阀门_开度反馈显示',
    'CV15_2_比例阀门_开度反馈显示',
    'CV2_1_比例阀门_开度反馈显示',
    'CV2_3_比例阀门_开度反馈显示',
    'CV9_1_比例阀门_开度反馈显示',
    'CV9_2_比例阀门_开度反馈显示',
    'V1_6_开关阀门_开到位反馈',
    'V1_7_开关阀门_开到位反馈',
    'V16_1_开关阀门_开到位反馈',
    'V16_2_开关阀门_开到位反馈',
    'CWP_1_冷却水泵_频率反馈显示'
]

# 加载工作簿
wb = load_workbook(updated_data_path)
ws = wb.active

# 创建红色字体样式
red_font = Font(color="FF0000")

# 1. 对比数据并标记差异
for col_idx, column in enumerate(columns_to_compare, start=2):  # Excel中的列从2开始
    if column in original_data.columns and column in updated_data.columns:
        original_column = original_data[column]
        updated_column = updated_data[column]

        for row_idx in range(len(original_column)):
            if row_idx < len(updated_column):  # 确保不会超出更新数据的范围
                if original_column[row_idx] != updated_column[row_idx]:
                    # 将差异单元格字体颜色改为红色
                    ws.cell(row=row_idx + 2, column=col_idx).font = red_font  # 行列从1开始，+2是因为数据有标题行

# 2. 添加新功能：比较能耗并标红
# 找到'优化后能耗'和'冷源群控系统实时能耗_单位_kW'的列位置
optimized_col = None
realtime_col = None

for col_idx in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col_idx).value
    if header == '优化后能耗':
        optimized_col = col_idx
    elif header == '冷源群控系统实时能耗_单位_kW':
        realtime_col = col_idx

# 如果找到这两列，进行比较
if optimized_col and realtime_col:
    for row_idx in range(2, ws.max_row + 1):  # 从第2行开始（跳过标题）
        optimized_value = ws.cell(row=row_idx, column=optimized_col).value
        realtime_value = ws.cell(row=row_idx, column=realtime_col).value

        # 确保两个值都是数字类型
        if isinstance(optimized_value, (int, float)) and isinstance(realtime_value, (int, float)):
            if optimized_value < realtime_value:
                # 标红优化后能耗列
                ws.cell(row=row_idx, column=optimized_col).font = red_font
            # 无论是否优化，这一行都会被保留

# 保存修改后的Excel文件（包含所有数据行）
updated_data_with_red_font_path = 'PSO调优后的参数_v2_标红.xlsx'
wb.save(updated_data_with_red_font_path)

print(f"处理完成，结果已保存到: {updated_data_with_red_font_path}")